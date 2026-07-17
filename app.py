import streamlit as st
import pandas as pd
import re
import io
import base64
import warnings
from collections import Counter
from datetime import datetime
from openpyxl.styles import PatternFill, Font

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

st.set_page_config(page_title="Validación de Documentos", layout="wide")

st.markdown("""
    <style>
    #MainMenu, footer, header { visibility: hidden; }
    .stApp { background-color: #f6f8fa; font-family: 'Segoe UI', Arial, sans-serif; }
    h1 { color: #0a3d62; border-left: 5px solid #DA291C; padding-left: 10px; }
    div.stButton > button { background-color: #DA291C !important; color: white !important; font-weight: bold !important; border-radius: 8px !important; }
    div.stDownloadButton > button { background-color: #1e293b !important; color: white !important; font-weight: bold !important; border-radius: 8px !important; }
    </style>
""", unsafe_allow_html=True)

st.title("📊 Validación y Cálculo de Primas - Seguros 📊")

with st.expander("ℹ️ Cómo funciona", expanded=False):
    st.markdown(
        "1. Elige la zona y tu usuario.\n"
        "2. Sube uno o más Excel con los registros a validar.\n"
        "3. Presiona **Procesar archivos** y descarga el reporte final."
    )

# ------------------ Selector de zona (afecta NETA) ------------------
# Las tasas son referenciales y configurables según la aseguradora

col_z1, col_z2, col_z3 = st.columns([1, 1, 1])
with col_z1:
    zona = st.selectbox("Selecciona la zona", options=["Sur", "Norte"], index=0)
with col_z2:
    usuarios = ["Sofi B", "Engel B", "User_01", "User_02"]
    usuario_seleccionado = st.selectbox("Selecciona tu usuario:", usuarios)
with col_z3:
    st.write("")
    st.write("")
    if "reset_id_primas" not in st.session_state:
        st.session_state.reset_id_primas = 0
    if st.button("🔄 Limpiar / Reiniciar", use_container_width=True):
        st.session_state.reset_id_primas += 1
        st.rerun()

NETA = 0.00038 if zona == "Sur" else 0.00036
V_D_E = 0.03
V_IGV = 0.18

# Fecha del reporte
fecha_reporte = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
st.caption(f"📅 Fecha del reporte: {fecha_reporte}")

# Subir archivos
archivos = st.file_uploader(
    "Sube tus archivos Excel", type=["xlsx"], accept_multiple_files=True,
    key=f"primas_uploader_{st.session_state.reset_id_primas}"
)

if not archivos:
    st.info("⬆️ Sube al menos un Excel para comenzar.")

if archivos:
    nombres = [a.name for a in archivos]
    duplicados = sorted({n for n, c in Counter(nombres).items() if c > 1})
    if duplicados:
        st.error(f"❌ Hay archivos duplicados, quítalos antes de continuar: {', '.join(duplicados)}")
        st.stop()

# Botón para procesar
if st.button("⚙️ Procesar archivos") and archivos:

    no_validos = []
    resumen = []
    errores_archivo = []  # [(nombre_archivo, mensaje_error)]

    def validar_documento(row):
        tipo = str(row.get("Tipo de Documento", "")).strip().upper()
        num = str(row.get("Número de Documento", "")).strip()
        if tipo == "DNI":
            return "DNI válido" if num.isdigit() and len(num) == 8 else "DNI inválido"
        return "No es DNI"

    progreso = st.progress(0, text="Procesando archivos...")
    total_archivos = len(archivos)

    for idx_archivo, archivo in enumerate(archivos, start=1):
        nombre_archivo = archivo.name
        try:
            df = pd.read_excel(archivo, dtype={"Número de Documento": str})
            df.columns = df.columns.str.strip()
            df = df.dropna(how="all")
            df["fila_en_excel"] = df.index + 2

            if df.empty:
                resumen.append({"Archivo": nombre_archivo, "Poliza": "no declara"})
                progreso.progress(idx_archivo / total_archivos, text=f"Procesando archivos... ({idx_archivo}/{total_archivos})")
                continue

            # Asegurar columnas obligatorias
            for col in ["Tipo de Documento", "Número de Documento", "Capital Asegurado", "Prima"]:
                if col not in df.columns:
                    df[col] = pd.NA
            if "Nombre Completo" not in df.columns:
                df["Nombre Completo"] = pd.NA

            # Validación DNI
            df["validación documento"] = df.apply(validar_documento, axis=1)

            # Filtrar no válidos
            df_no_validos = df[df["validación documento"] == "No es DNI"].copy()
            df_no_validos["archivo_origen"] = nombre_archivo

            columnas_finales = [
                "Tipo de Documento", "Número de Documento", "Nombre Completo",
                "validación documento", "archivo_origen", "fila_en_excel"
            ]
            for col in columnas_finales:
                if col not in df_no_validos.columns:
                    df_no_validos[col] = pd.NA
            df_no_validos = df_no_validos[columnas_finales]

            df_no_validos = df_no_validos[
                df_no_validos["Número de Documento"].notna() &
                df_no_validos["Número de Documento"].astype(str).str.strip().ne("") &
                df_no_validos["Nombre Completo"].notna() &
                df_no_validos["Nombre Completo"].astype(str).str.strip().ne("")
            ]

            if not df_no_validos.empty:
                no_validos.append(df_no_validos)

            # Detectar si última fila es TOTAL
            ultima_es_subtotal = df.iloc[-1].astype(str).str.contains("TOTAL", case=False, na=False).any()

            if ultima_es_subtotal and len(df) > 1:
                ultima_fila = df.iloc[-1]
                df_sin_ultima = df.iloc[:-1].copy()
                sub_capital = ultima_fila.get("Capital Asegurado", "no declara")
                sub_prima = ultima_fila.get("Prima", "no declara")
            else:
                df_sin_ultima = df.copy()
                sub_capital = "no declara"
                sub_prima = "no declara"

            # Totales existentes
            total_capital_num = df_sin_ultima["Capital Asegurado"].sum(min_count=1)

            s = (df_sin_ultima["Prima"].astype(str)
                 .str.replace('\u00A0', '', regex=False)
                 .str.replace('\u202F', '', regex=False)
                 .str.replace(' ', '', regex=False)
                 .str.replace('S/', '', regex=False)
                 .str.replace('s/', '', regex=False)
                 .str.replace('.', '', regex=False)
                 .str.replace(',', '.', regex=False))

            total_prima_num = pd.to_numeric(s, errors="coerce").sum(min_count=1)

            # ---------- Cálculos ----------
            capital_num = pd.to_numeric(df_sin_ultima["Capital Asegurado"], errors="coerce")

            prima_neta_reg = capital_num * NETA
            d_e_reg = prima_neta_reg * V_D_E
            igv_reg = (prima_neta_reg + d_e_reg) * V_IGV
            total_reg = prima_neta_reg + d_e_reg + igv_reg

            def red2(x):
                return float(round(x, 2)) if pd.notna(x) else "no declara"

            suma_prima_neta = red2(prima_neta_reg.sum(min_count=1))
            suma_d_e = red2(d_e_reg.sum(min_count=1))
            suma_igv = red2(igv_reg.sum(min_count=1))
            suma_total = red2(total_reg.sum(min_count=1))

            # Extraer póliza
            match = re.search(r'\d{10,}', nombre_archivo)
            poliza = match.group(0) if match else "no declara"

            # ---------- RESUMEN ----------
            resumen.append({
                "Archivo": nombre_archivo,
                "Poliza": poliza,
                "Usuario": usuario_seleccionado,
                "Zona": zona,
                "Fecha_reporte": fecha_reporte,
                "Cantidad_registros": len(df_sin_ultima),
                "Total_capital": total_capital_num,
                "Total_origen_col_H": sub_capital,
                "Total_origen_col_J": sub_prima,
                "prima_neta": suma_prima_neta,
                "D_E": suma_d_e,
                "IGV": suma_igv,
                "TOTAL": suma_total
            })
        except Exception as e:
            errores_archivo.append((nombre_archivo, str(e)))

        progreso.progress(idx_archivo / total_archivos, text=f"Procesando archivos... ({idx_archivo}/{total_archivos})")

    progreso.empty()

    df_no_validos_final = pd.concat(no_validos, ignore_index=True) if no_validos else pd.DataFrame()
    df_resumen = pd.DataFrame(resumen)

    # Orden columnas
    orden_cols = [
        "Archivo", "Poliza", "Usuario", "Zona", "Fecha_reporte",
        "Cantidad_registros", "Total_capital",
        "Total_origen_col_H", "Total_origen_col_J",
        "prima_neta", "D_E", "IGV", "TOTAL"
    ]
    for col in orden_cols:
        if col not in df_resumen.columns:
            df_resumen[col] = pd.NA
    df_resumen = df_resumen[orden_cols]

    st.success("✅ Proceso completado.")

    m1, m2, m3 = st.columns(3)
    m1.metric("Archivos procesados", total_archivos)
    m2.metric("Registros totales", int(df_resumen["Cantidad_registros"].fillna(0).sum()) if not df_resumen.empty else 0)
    m3.metric("Documentos no válidos", len(df_no_validos_final))

    if errores_archivo:
        with st.expander(f"❌ {len(errores_archivo)} archivo(s) con error al procesar", expanded=True):
            for nombre, msg in errores_archivo:
                st.error(f"{nombre}: {msg}")

    with st.expander(f"📊 Totales por archivo (mostrando 5 de {len(df_resumen)})", expanded=False):
        st.dataframe(df_resumen.head(5), use_container_width=True)

    with st.expander(f"⚠️ No válidos (mostrando 5 de {len(df_no_validos_final)})", expanded=False):
        st.dataframe(df_no_validos_final.head(5), use_container_width=True)

    # Exportar a Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # PRIMERO: Totales por archivo
        df_resumen.to_excel(writer, sheet_name="Totales por archivo", index=False)

        # SEGUNDO: No válidos
        df_no_validos_final.to_excel(writer, sheet_name="No válidos", index=False)

        # ----------- COLOR A LAS CABECERAS -----------
        wb = writer.book
        fill = PatternFill(start_color="D53032", end_color="D53032", fill_type="solid")
        font_white = Font(color="FFFFFF", bold=True)

        hojas = ["Totales por archivo", "No válidos"]

        for hoja in hojas:
            ws = wb[hoja]
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font_white

    sello_fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_excel = f"Resumen_Validacion_{sello_fecha}.xlsx"

    # Descarga vía link base64 (evita el problema de ZIP/Excel truncado por proxy/antivirus corporativo)
    b64_excel = base64.b64encode(output.getvalue()).decode()
    href_excel = (
        f'<a download="{nombre_excel}" '
        f'href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_excel}" '
        f'style="display:inline-block;padding:0.6em 1.4em;background-color:#1e293b;color:white;'
        f'font-weight:bold;border-radius:8px;text-decoration:none;">📥 Descargar reporte final</a>'
    )
    st.markdown(href_excel, unsafe_allow_html=True)
